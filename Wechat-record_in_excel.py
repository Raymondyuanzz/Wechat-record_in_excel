#coding=utf-8
#-------------------------------------------------------------------------------
# Name:		
# Purpose:	获取YSH同学群里同学们发的文件，汇总为表格，用于美文阅读统计。
#Version： A01
# Author:	   YZZ，YSH
# Created:	2019-05-09
#-------------------------------------------------------------------------------

import os
import sys
import time
import datetime

import itchat
from itchat.content import *

import openpyxl
import openpyxl.styles as sty
from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter

group_name1 = "家情聊"
group_name2 = "2017级"
excel_filename = "87List.xlsx"
name_count = 48
time_run     = "22:05:00"             # 什么时候执行
time_end    = "22:10:00"             # 终止执行时间


def TimeStampToTime(timestamp):
	'''把时间戳转化为时间: 1479264792 to 2016-11-16 10:53:12'''
	timeStruct = time.localtime(timestamp)
	return time.strftime('%Y-%m-%d %H:%M:%S',timeStruct)
	#return time.strftime('%H:%M:%S',timeStruct)	
def get_FileSize(filePath):
	'''获取文件的大小,结果保留两位小数，单位为MB'''
	filePath = unicode(filePath,'utf8')
	fsize = os.path.getsize(filePath)
	fsize = fsize/float(1024*1024)
	return round(fsize,2)
def get_FileCreateTime(filePath):
	'''获取文件的创建时间'''
	filePath = unicode(filePath,'utf8')
	t = os.path.getctime(filePath)
	return TimeStampToTime(t)


def reply():
	#群文件保存
	@itchat.msg_register(['Picture','Recording', 'Attachment', 'Video'], isGroupChat=True )

	def download_files(msg):
		#print (msg)
		send_flag  = 1
		g_Sender     = msg["ActualNickName"]
		g_FileName = msg["FileName"]
		g_CreateTime = TimeStampToTime(msg["CreateTime"])
		chatroom_name = msg["User"]["NickName"]
		if (msg["FileSize"] != ''):
			g_FileSize    = round(int(msg["FileSize"])/float(1024),2)    #KB  
		else:
			g_FileSize = 1
		
		print ("Sender: ", g_Sender)
		print (g_FileName)
		print (g_FileSize, "KB")
		print (g_CreateTime)
		print ("Group: ", chatroom_name)
		
		if ("发送汇总表格给我" in g_FileName):
			send_flag = 11
		
		#建立Folder保存文件
		path = os.path.join(os.path.abspath("."),"group_files")
		file_name_list = os.listdir(path)
		chatroom_file_location = os.path.join(os.path.abspath("."),"group_files","%s" % (chatroom_name))
		if (chatroom_name not in file_name_list):
			os.mkdir(chatroom_file_location)
		#save_path = os.path.join(chatroom_file_location, "YZZ"+msg["FileName"])
		save_path = os.path.join(chatroom_file_location, str(msg['ActualNickName']) +"_"+ msg["FileName"])
		msg['Text'](save_path)
		
		#记录信息到Excel文件
		if (chatroom_name==group_name2): # or (chatroom_name==group_name1):
			#record in excel file.
			Datestr = datetime.datetime.now().strftime("%Y-%m-%d")			
			wb = Workbook()
			wb = load_workbook(filename = excel_filename )
			sheetnames = wb.worksheets
			ws = sheetnames[len(wb.worksheets)-1] #在最后一个sheet页操作
			print ("Write in sheet :", ws.title)
			if ws.title != Datestr:
				print (ws.title," != ", Datestr)
				ws = wb.copy_worksheet(wb['Data'])
				ws.title = Datestr
				#ws = wb.active
				#ws = wb.create_sheet(title = Datestr,index =0) #可传title和index两个参数,不传生成的WorkSheet名在'Sheet'后面递增加数字
				print (ws)
			
			#write info in excel cell
			w_flag = 0
			for i in range(2, name_count):
				Name = ws.cell(i, 1).value
				eNickName = Name
				if (ws.cell(i, 2).value is not None):
					eNickName = ws.cell(i, 2).value
				#print (ws.cell(i,1).value)
				if (Name is not None) and ((Name in g_Sender) or (g_Sender in Name) or (Name in g_FileName) or (eNickName in g_Sender) or (g_Sender in eNickName) or (eNickName in g_FileName)):
					if (ws.cell(i, 3).value is None): #未填写文件名
						print ("Write", "--> ", Name)
						ws.cell(i, 3).value= g_FileName
						ws.cell(i, 4).value= g_FileSize
						ws.cell(i, 5).value= g_CreateTime			
						ws.cell(i, 1).fill=sty.PatternFill(fill_type='solid',fgColor="43CD80")  #http://www.114la.com/other/rgb.htm
					else:                                      #重复发送文件
						for k in range(7, 100):
							if (ws.cell(i, k).value is None):
								ws.cell(i, k).value= g_FileName
								ws.cell(i, 1).fill=sty.PatternFill(fill_type='solid',fgColor="FF83FA") #重复发送则变更颜色
								w_flag = 1
								break
						print ("Duplicated", "--> ", g_Sender)
					w_flag = 1
					break 

			if (w_flag == 0):                             #非表中人员发送，则独立记录
				for jk in range(50, 9999):
					if (ws.cell(jk, 1).value is None):
						ws.cell(jk, 1).value = g_Sender
						ws.cell(jk, 3).value= g_FileName
						ws.cell(jk, 4).value= g_FileSize
						ws.cell(jk, 5).value= g_CreateTime						
						#ws.cell(jk, 5).value= str(msg)
						print ("J = ", "--> ", jk-49)
						break 
			wb.save(excel_filename)

		#发送表格给特定人员
		now_stamp  = time.strftime("%H:%M:%S", time.localtime())
		if (((now_stamp > time_run) and (now_stamp < time_end) and send_flag == 1) or (send_flag == 11)):
			print ("Send file ", "--> ", "Excel file!")
			#print ("File Helper: ", itchat.send('Hello, filehelper', toUserName='filehelper'))
			friends_list = itchat.get_friends(update=True)    #想给谁发信息，先查找到这个朋友
			users1 = itchat.search_friends(name=u'元亨利贞') #找到UserName   
			#users = itchat.search_friends(name=u'千寻之恋') #找到UserName    
			users2 = itchat.search_friends(name=u'Eileen')    #找到UserName    
			userName1 = users1[0]['UserName']
			userName2 = users2[0]['UserName']			
			#print (userNamex)
			#print ("Send msg result: ", itchat.send(msg="Text Message", toUserName = userNamex))
			print ("Send msg result: ", itchat.send('@fil@%s' % 'E:\\Downloads\\wechatRoom-master\\87List.xlsx', toUserName=userName1))
			print ("Send msg result: ", itchat.send('@fil@%s' % 'E:\\Downloads\\wechatRoom-master\\87List.xlsx', toUserName=userName2))
			send_flag = 0


if __name__ == '__main__':
	reply()
	itchat.auto_login(hotReload=True)  #加上hotReload==True,那么就会保留登录的状态，至少在后面的几次登录过程中不会再次扫描二维码，该参数生成一个静态文件itchat.pkl用于存储登录状态
	itchat.run()
